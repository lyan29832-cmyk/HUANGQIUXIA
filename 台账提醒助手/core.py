from __future__ import annotations

import hashlib
import json
import smtplib
import ssl
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from email.header import Header
from email.mime.text import MIMEText
from email.utils import formataddr
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

IMPACT_VALUE = "直接影响"
REQUIRED_HEADERS = ("事项编号", "事项名称", "影响程度", "责任人", "责任人邮箱")
OPTIONAL_HEADERS = ("企微账号",)

SMTP_PRESETS: dict[str, tuple[str, int]] = {
    "QQ邮箱": ("smtp.qq.com", 465),
    "163邮箱": ("smtp.163.com", 465),
    "126邮箱": ("smtp.126.com", 465),
    "腾讯企业邮箱": ("smtp.exmail.qq.com", 465),
    "阿里企业邮箱": ("smtp.qiye.aliyun.com", 465),
    "网易企业邮箱": ("smtp.qiye.163.com", 465),
    "Outlook / 微软": ("smtp.office365.com", 587),
}


@dataclass
class AppConfig:
    excel_path: str = ""
    sheet_name: str = ""
    interval_minutes: int = 10
    enable_email: bool = True
    enable_wecom: bool = False
    mail_preset: str = "QQ邮箱"
    smtp_host: str = "smtp.qq.com"
    smtp_port: int = 465
    mail_user: str = ""
    mail_pass: str = ""
    mail_from_name: str = "台账提醒助手"
    wecom_webhook: str = ""


@dataclass
class LedgerRow:
    row_key: str
    item_id: str
    title: str
    impact: str
    owner: str
    email: str
    wecom: str
    row_number: int


@dataclass
class CheckResult:
    file_missing: bool = False
    not_updated_today: bool = False
    first_run: bool = False
    missing_headers: list[str] = field(default_factory=list)
    hits: list[LedgerRow] = field(default_factory=list)
    message: str = ""
    notified: int = 0
    errors: list[str] = field(default_factory=list)


def config_path(base_dir: Path) -> Path:
    return base_dir / "config.json"


def data_dir(base_dir: Path) -> Path:
    path = base_dir / "data"
    path.mkdir(parents=True, exist_ok=True)
    return path


def load_config(base_dir: Path) -> AppConfig:
    path = config_path(base_dir)
    if not path.exists():
        return AppConfig()
    raw = json.loads(path.read_text(encoding="utf-8"))
    allowed = {k: raw[k] for k in AppConfig.__dataclass_fields__ if k in raw}
    return AppConfig(**allowed)


def save_config(base_dir: Path, config: AppConfig) -> None:
    config_path(base_dir).write_text(
        json.dumps(asdict(config), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def snapshot_path(base_dir: Path, excel_path: str) -> Path:
    digest = hashlib.sha256(excel_path.encode("utf-8")).hexdigest()[:16]
    return data_dir(base_dir) / f"snapshot-{digest}.json"


def infer_smtp(email_addr: str, preset: str) -> tuple[str, int]:
    if preset in SMTP_PRESETS:
        return SMTP_PRESETS[preset]
    domain = email_addr.split("@")[-1].lower() if "@" in email_addr else ""
    mapping = {
        "qq.com": ("smtp.qq.com", 465),
        "163.com": ("smtp.163.com", 465),
        "126.com": ("smtp.126.com", 465),
        "exmail.qq.com": ("smtp.exmail.qq.com", 465),
        "outlook.com": ("smtp.office365.com", 587),
        "hotmail.com": ("smtp.office365.com", 587),
        "live.com": ("smtp.office365.com", 587),
    }
    return mapping.get(domain, ("smtp.qq.com", 465))


def file_updated_today(path: Path, today: date | None = None) -> bool:
    today = today or date.today()
    return datetime.fromtimestamp(path.stat().st_mtime).date() == today


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_map(ws: Worksheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in ws[1]:
        name = _cell_text(cell.value)
        if name:
            mapping[name] = cell.column
    return mapping


def read_ledger_rows(excel_path: Path, sheet_name: str = "") -> tuple[list[LedgerRow], list[str], str]:
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        used_sheet = sheet_name if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames[0]
        ws = wb[used_sheet]
        headers = _header_map(ws)
        missing = [name for name in REQUIRED_HEADERS if name not in headers]
        if missing:
            return [], missing, used_sheet

        rows: list[LedgerRow] = []
        for idx, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

            def col(name: str) -> str:
                pos = headers.get(name)
                if not pos:
                    return ""
                value = raw[pos - 1] if pos - 1 < len(raw) else None
                return _cell_text(value)

            if not any(_cell_text(v) for v in raw or ()):
                continue
            item_id = col("事项编号")
            title = col("事项名称")
            row_key = item_id or title or f"row-{idx}"
            rows.append(
                LedgerRow(
                    row_key=row_key,
                    item_id=item_id,
                    title=title,
                    impact=col("影响程度"),
                    owner=col("责任人"),
                    email=col("责任人邮箱"),
                    wecom=col("企微账号"),
                    row_number=idx,
                )
            )
        return rows, [], used_sheet
    finally:
        wb.close()


def list_sheet_names(excel_path: Path) -> list[str]:
    wb = load_workbook(excel_path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def rows_to_snapshot(rows: list[LedgerRow]) -> dict[str, dict[str, str]]:
    return {
        row.row_key: {
            "impact": row.impact,
            "item_id": row.item_id,
            "title": row.title,
        }
        for row in rows
    }


def load_snapshot(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def save_snapshot(path: Path, rows: list[LedgerRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(rows_to_snapshot(rows), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def find_new_direct_impacts(
    old: dict[str, dict[str, str]],
    rows: list[LedgerRow],
    target: str = IMPACT_VALUE,
) -> list[LedgerRow]:
    hits: list[LedgerRow] = []
    for row in rows:
        if row.impact != target:
            continue
        previous = old.get(row.row_key, {})
        if previous.get("impact") != target:
            hits.append(row)
    return hits


def build_message(row: LedgerRow) -> tuple[str, str]:
    subject = f"【台账预警】{row.item_id or row.title} 影响程度为{IMPACT_VALUE}"
    body = (
        "台账有更新，需要关注：\n"
        f"事项编号：{row.item_id or '（空）'}\n"
        f"事项名称：{row.title or '（空）'}\n"
        f"责任人：{row.owner or '（空）'}\n"
        f"影响程度已改为：{IMPACT_VALUE}\n"
    )
    return subject, body


def send_email(config: AppConfig, to_addr: str, subject: str, body: str) -> None:
    if not to_addr:
        raise ValueError("这一行没有填写责任人邮箱")
    host, port = config.smtp_host, int(config.smtp_port)
    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = Header(subject, "utf-8")
    msg["From"] = formataddr((config.mail_from_name, config.mail_user))
    msg["To"] = to_addr

    if port == 465:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(host, port, context=context, timeout=20) as smtp:
            smtp.login(config.mail_user, config.mail_pass)
            smtp.sendmail(config.mail_user, [to_addr], msg.as_string())
        return

    with smtplib.SMTP(host, port, timeout=20) as smtp:
        smtp.starttls(context=ssl.create_default_context())
        smtp.login(config.mail_user, config.mail_pass)
        smtp.sendmail(config.mail_user, [to_addr], msg.as_string())


def send_wecom_bot(webhook: str, content: str, mentioned: list[str] | None = None) -> dict[str, Any]:
    if not webhook:
        raise ValueError("还没有填写企业微信机器人地址")
    payload = {
        "msgtype": "text",
        "text": {
            "content": content,
            "mentioned_list": mentioned or [],
        },
    }
    response = requests.post(webhook, json=payload, timeout=15)
    response.raise_for_status()
    data = response.json()
    if data.get("errcode") not in (0, None):
        raise RuntimeError(f"企业微信返回错误：{data}")
    return data


def notify_hits(config: AppConfig, hits: list[LedgerRow]) -> tuple[int, list[str]]:
    sent = 0
    errors: list[str] = []
    for row in hits:
        subject, body = build_message(row)
        try:
            if config.enable_email:
                send_email(config, row.email, subject, body)
            if config.enable_wecom:
                mentioned = [row.wecom] if row.wecom else []
                send_wecom_bot(config.wecom_webhook, f"{subject}\n{body}", mentioned)
            sent += 1
        except Exception as exc:  # noqa: BLE001 — show each row's failure in the UI log
            errors.append(f"{row.item_id or row.title}: {exc}")
    return sent, errors


def check_and_notify(
    base_dir: Path,
    config: AppConfig,
    *,
    ignore_mtime: bool = False,
    today: date | None = None,
) -> CheckResult:
    result = CheckResult()
    excel_path = Path(config.excel_path)
    if not config.excel_path or not excel_path.exists():
        result.file_missing = True
        result.message = "找不到台账文件。请先点「选择台账文件」，选爱数同步盘里的 Excel。"
        return result

    snap = snapshot_path(base_dir, str(excel_path.resolve()))
    try:
        rows, missing, used_sheet = read_ledger_rows(excel_path, config.sheet_name)
    except Exception as exc:  # noqa: BLE001
        result.message = f"打不开 Excel，请确认文件没有损坏，且爱数已经同步完成。详情：{exc}"
        return result

    if missing:
        result.missing_headers = missing
        result.message = (
            "Excel 第 1 行缺少这些表头："
            + "、".join(missing)
            + "。请改成：事项编号、事项名称、影响程度、责任人、责任人邮箱。"
        )
        return result

    is_first = not snap.exists()
    old = {} if is_first else load_snapshot(snap)

    if is_first:
        save_snapshot(snap, rows)
        result.first_run = True
        result.message = (
            f"第一次检查完成（工作表：{used_sheet}）。"
            "已记住当前内容，不会发通知。以后有人把「影响程度」改成「直接影响」才会提醒。"
        )
        return result

    if not ignore_mtime and not file_updated_today(excel_path, today=today):
        result.not_updated_today = True
        result.message = "台账文件今天还没有保存过新内容，无需通知。"
        save_snapshot(snap, rows)
        return result

    hits = find_new_direct_impacts(old, rows)
    result.hits = hits
    if not hits:
        save_snapshot(snap, rows)
        result.message = "检查完成：今天有更新，但没有新增「直接影响」。"
        return result

    sent, errors = notify_hits(config, hits)
    result.notified = sent
    result.errors = errors
    save_snapshot(snap, rows)
    names = "、".join(h.item_id or h.title or f"第{h.row_number}行" for h in hits)
    result.message = f"发现 {len(hits)} 条新的「直接影响」：{names}。已尝试通知 {sent} 条。"
    return result


def create_template(save_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "台账"

    headers = list(REQUIRED_HEADERS) + list(OPTIONAL_HEADERS)
    fills = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(1, col, name)
        cell.fill = fills
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    samples = [
        ("001", "示例：服务器故障", "间接影响", "张三", "zhangsan@example.com", ""),
        ("002", "示例：客户投诉", "无影响", "李四", "lisi@example.com", ""),
    ]
    for r, sample in enumerate(samples, start=2):
        for c, value in enumerate(sample, start=1):
            ws.cell(r, c, value)

    dv = DataValidation(type="list", formula1='"直接影响,间接影响,无影响"', allow_blank=True)
    dv.error = "请从下拉中选择"
    dv.errorTitle = "输入无效"
    dv.prompt = "请选择影响程度"
    dv.promptTitle = "影响程度"
    ws.add_data_validation(dv)
    dv.add("C2:C500")

    widths = {"A": 14, "B": 28, "C": 14, "D": 12, "E": 28, "F": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    save_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(save_path)
    return save_path
